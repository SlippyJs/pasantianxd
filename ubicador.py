"""
ubicador.py
===========
Responsabilidad ÚNICA: recibir las filas extraídas, enriquecerlas con
información de ubicación (archivo PDF, página, línea) y generar el Excel
final con todo en un solo archivo organizado.

Genera un Excel con DOS hojas:
  • "Datos Extraídos"  → todos los registros con sus columnas de ubicación
  • "Resumen por PDF"  → cuántas filas, qué páginas por cada archivo PDF
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Definición de columnas del Excel ────────────────────────────────────────
# (nombre_interno, encabezado_visible, ancho_columna, es_columna_ubicacion)

COLUMNAS_EXCEL = [
    ("numero_item",    "Ítem",           9,   False),
    ("cantidad",       "Cantidad",       11,  False),
    ("descripcion",    "Descripción",    44,  False),
    ("valor_unitario", "Valor Unitario", 16,  False),
    ("valor_total",    "Valor Total",    16,  False),
    ("archivo_origen", "Archivo PDF",    30,  True),
    ("numero_pagina",  "Página",         10,  True),
    ("numero_linea",   "Línea",          10,  True),
]

# ─── Paleta de colores ────────────────────────────────────────────────────────

_COLOR_HEADER_DATOS    = "1F4E79"   # Azul oscuro  → encabezados de datos
_COLOR_HEADER_UBIC     = "7B3F00"   # Café oscuro  → encabezados de ubicación
_COLOR_TEXTO_HEADER    = "FFFFFF"   # Blanco para texto de encabezados
_COLOR_FILA_PAR_DATOS  = "D6E4F0"   # Azul claro
_COLOR_FILA_IMPAR_DATOS= "FFFFFF"   # Blanco
_COLOR_FILA_PAR_UBIC   = "FFF3CD"   # Amarillo claro
_COLOR_FILA_IMPAR_UBIC = "FFFDE7"   # Amarillo muy claro

_BORDE = Border(
    left   = Side(style="thin", color="BFBFBF"),
    right  = Side(style="thin", color="BFBFBF"),
    top    = Side(style="thin", color="BFBFBF"),
    bottom = Side(style="thin", color="BFBFBF"),
)


# ─── Clase pública ───────────────────────────────────────────────────────────

class UbicadorYGeneradorExcel:
    """
    Recibe la lista de filas extraídas (ya con numero_pagina y numero_linea),
    les agrega el nombre del archivo de origen y genera el Excel final.

    Uso:
        ubicador = UbicadorYGeneradorExcel(ruta_carpeta_salida)
        ubicador.agregar_filas(filas_del_extractor, nombre_pdf)
        ruta = ubicador.generar_excel()
    """

    def __init__(self, ruta_carpeta_salida: str):
        self.ruta_carpeta_salida = ruta_carpeta_salida
        self._todas_las_filas: list = []

    # ── Agregar filas con nombre del archivo de origen ────────────────────────

    def agregar_filas(self, filas: list, nombre_archivo_pdf: str) -> None:
        """
        Toma las filas del extractor y les añade la columna 'archivo_origen'.
        Llamar una vez por cada PDF procesado.
        Asegura que todos los valores sean strings válidos (nunca None).
        """
        for fila in filas:
            fila_enriquecida = dict(fila)
            # Garantizar que no hay valores None
            for clave in ["numero_item", "cantidad", "descripcion", "valor_unitario", "valor_total"]:
                if clave not in fila_enriquecida:
                    fila_enriquecida[clave] = ""
                elif fila_enriquecida[clave] is None:
                    fila_enriquecida[clave] = ""
            fila_enriquecida["archivo_origen"] = nombre_archivo_pdf
            self._todas_las_filas.append(fila_enriquecida)

    def total_filas(self) -> int:
        return len(self._todas_las_filas)

    # ── Generación del Excel ──────────────────────────────────────────────────

    def generar_excel(self, nombre_archivo: str = "resultado_extraccion.xlsx") -> str:
        """
        Genera el Excel con las hojas 'Datos Extraídos' y 'Resumen por PDF'.
        Retorna la ruta completa del archivo generado.
        """
        if not self._todas_las_filas:
            raise ValueError("No hay filas para exportar al Excel.")

        tabla = pd.DataFrame(self._todas_las_filas)

        # Asegurar que todas las columnas existan (aunque estén vacías)
        for nombre_col, *_ in COLUMNAS_EXCEL:
            if nombre_col not in tabla.columns:
                tabla[nombre_col] = ""

        # Resolver nombre de archivo sin sobreescribir
        ruta_salida = self._resolver_ruta(nombre_archivo)

        libro = Workbook()

        # Hoja 1: Datos + ubicación
        hoja_datos       = libro.active
        hoja_datos.title = "Datos Extraídos"
        self._escribir_hoja_datos(hoja_datos, tabla)

        # Hoja 2: Resumen por PDF
        hoja_resumen = libro.create_sheet("Resumen por PDF")
        self._escribir_hoja_resumen(hoja_resumen, tabla)

        libro.save(ruta_salida)
        return ruta_salida

    # ── Escritura de la hoja de datos ─────────────────────────────────────────

    def _escribir_hoja_datos(self, hoja, tabla: pd.DataFrame) -> None:
        """Escribe todas las filas con sus columnas de ubicación."""

        # ── Fila 1: Encabezados ──────────────────────────────────────────────
        for col_idx, (nombre, encabezado, ancho, es_ubic) in enumerate(COLUMNAS_EXCEL, start=1):
            celda = hoja.cell(row=1, column=col_idx, value=encabezado)
            color = _COLOR_HEADER_UBIC if es_ubic else _COLOR_HEADER_DATOS
            celda.fill      = PatternFill("solid", fgColor=color)
            celda.font      = Font(color=_COLOR_TEXTO_HEADER, bold=True, size=10)
            celda.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
            celda.border    = _BORDE
            hoja.column_dimensions[get_column_letter(col_idx)].width = ancho

        hoja.row_dimensions[1].height = 32

        # ── Filas de datos ───────────────────────────────────────────────────
        for fila_idx, (_, fila_df) in enumerate(tabla.iterrows(), start=2):
            es_par = (fila_idx % 2 == 0)

            for col_idx, (nombre, _, _, es_ubic) in enumerate(COLUMNAS_EXCEL, start=1):
                valor = fila_df.get(nombre, "")
                # Garantizar que valor nunca es None o NaN
                if valor is None or (isinstance(valor, float) and pd.isna(valor)):
                    valor = ""
                else:
                    valor = str(valor).strip() if valor else ""
                
                celda = hoja.cell(row=fila_idx, column=col_idx, value=valor)

                # Color de fondo según tipo de columna y paridad
                if es_ubic:
                    color_fondo = _COLOR_FILA_PAR_UBIC if es_par else _COLOR_FILA_IMPAR_UBIC
                else:
                    color_fondo = _COLOR_FILA_PAR_DATOS if es_par else _COLOR_FILA_IMPAR_DATOS

                celda.fill   = PatternFill("solid", fgColor=color_fondo)
                celda.border = _BORDE
                celda.font   = Font(size=9)

                # Centrar columnas numéricas
                centrar = nombre in ("numero_item", "cantidad", "valor_unitario",
                                     "valor_total", "numero_pagina", "numero_linea")
                celda.alignment = Alignment(
                    horizontal="center" if centrar else "left",
                    vertical="center",
                    wrap_text=(nombre == "descripcion")
                )

        # Inmovilizar encabezado y activar autofiltros
        hoja.freeze_panes = "A2"
        ultima_col        = get_column_letter(len(COLUMNAS_EXCEL))
        hoja.auto_filter.ref = f"A1:{ultima_col}1"

    # ── Escritura de la hoja de resumen ───────────────────────────────────────

    def _escribir_hoja_resumen(self, hoja, tabla: pd.DataFrame) -> None:
        """Resumen por archivo PDF: filas extraídas y páginas con datos."""

        encabezados = [
            "Archivo PDF", "Filas Extraídas",
            "Páginas con Datos", "Pág. Mínima", "Pág. Máxima"
        ]
        anchos = [32, 16, 34, 13, 13]

        # ── Encabezados ──────────────────────────────────────────────────────
        for col_idx, (enc, ancho) in enumerate(zip(encabezados, anchos), start=1):
            celda = hoja.cell(row=1, column=col_idx, value=enc)
            color = _COLOR_HEADER_UBIC if col_idx > 1 else _COLOR_HEADER_DATOS
            celda.fill      = PatternFill("solid", fgColor=color)
            celda.font      = Font(color=_COLOR_TEXTO_HEADER, bold=True, size=10)
            celda.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
            celda.border    = _BORDE
            hoja.column_dimensions[get_column_letter(col_idx)].width = ancho

        hoja.row_dimensions[1].height = 28

        # ── Calcular resumen ─────────────────────────────────────────────────
        if "archivo_origen" not in tabla.columns or "numero_pagina" not in tabla.columns:
            return

        # Convertir numero_pagina a entero para ordenar correctamente
        tabla_resumen = tabla.copy()
        tabla_resumen["numero_pagina"] = pd.to_numeric(
            tabla_resumen["numero_pagina"], errors="coerce"
        )

        def agg_paginas(x):
            """Crea lista única de páginas sin None/NaN."""
            paginas = x.dropna().astype(int).unique()
            if len(paginas) == 0:
                return ""
            return ", ".join(str(p) for p in sorted(paginas))

        resumen = tabla_resumen.groupby("archivo_origen").agg(
            filas           = ("descripcion", "count"),
            paginas_unicas  = ("numero_pagina", agg_paginas),
            pag_min         = ("numero_pagina", "min"),
            pag_max         = ("numero_pagina", "max"),
        ).reset_index()

        # ── Filas de resumen ─────────────────────────────────────────────────
        for fila_idx, (_, fila) in enumerate(resumen.iterrows(), start=2):
            es_par = (fila_idx % 2 == 0)
            valores = [
                str(fila.get("archivo_origen", "") or ""),
                int(fila.get("filas", 0) or 0),
                str(fila.get("paginas_unicas", "") or ""),
                int(fila.get("pag_min", 0) or 0) if pd.notna(fila.get("pag_min")) else "",
                int(fila.get("pag_max", 0) or 0) if pd.notna(fila.get("pag_max")) else "",
            ]
            for col_idx, valor in enumerate(valores, start=1):
                celda = hoja.cell(row=fila_idx, column=col_idx, value=valor)
                es_ubic = col_idx > 1
                if es_ubic:
                    color = _COLOR_FILA_PAR_UBIC if es_par else _COLOR_FILA_IMPAR_UBIC
                else:
                    color = _COLOR_FILA_PAR_DATOS if es_par else _COLOR_FILA_IMPAR_DATOS
                celda.fill      = PatternFill("solid", fgColor=color)
                celda.border    = _BORDE
                celda.font      = Font(size=9)
                celda.alignment = Alignment(
                    horizontal="center" if col_idx > 1 else "left",
                    vertical="center"
                )

        hoja.freeze_panes = "A2"

    # ── Utilidades ────────────────────────────────────────────────────────────

    def _resolver_ruta(self, nombre_archivo: str) -> str:
        """Genera ruta de salida sin sobreescribir archivos existentes."""
        ruta = os.path.join(self.ruta_carpeta_salida, nombre_archivo)
        contador = 1
        nombre_base, ext = os.path.splitext(nombre_archivo)
        while os.path.exists(ruta):
            ruta = os.path.join(self.ruta_carpeta_salida,
                                f"{nombre_base}_{contador}{ext}")
            contador += 1
        return ruta
